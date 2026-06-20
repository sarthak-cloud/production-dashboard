import os
import re
from datetime import datetime
from openpyxl import load_workbook

DATABASE_ROOT = r"C:\Users\IR1035\Desktop\Database"


# =====================================================
# HELPERS
# =====================================================

def safe_str(value):
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
    return str(value).strip()


def is_valid_mid(value):
    """
    A valid MID looks like:  FHD-H01F-0080M  or  ABC-1234-5678X
    Rules:
      - 2 to 5 segments separated by hyphens
      - Each segment is short (1–10 chars), alphanumeric only
      - Total length between 5 and 40 characters
      - Must contain at least one hyphen
    """
    if not value:
        return False
    text = str(value).strip()
    if len(text) < 5 or len(text) > 40:
        return False
    if "-" not in text:
        return False
    if not re.fullmatch(r"[A-Za-z0-9]+(-[A-Za-z0-9]+){1,4}", text):
        return False
    return True


def normalize_shift(raw):
    """
    Convert any shift representation to "1", "2", "3", or "".
    Handles: int, float, ordinals, labels, named A/B/C.
    FIX: always return a string ("1"/"2"/"3"/"") for consistent
         comparisons — the DB stores shifts as integers (1/2/3)
         written by consolidator, so cast those first.
    """
    if raw is None:
        return ""

    # Numeric types — DB stores 1/2/3 as int
    if isinstance(raw, (int, float)):
        v = int(raw)
        if v in (1, 2, 3):
            return str(v)
        return ""

    text = str(raw).strip().upper()

    if text in ("1", "2", "3"):
        return text

    # Ordinal suffixes  1ST / 2ND / 3RD
    if text.startswith("1"): return "1"
    if text.startswith("2"): return "2"
    if text.startswith("3"): return "3"

    # Named shifts A/B/C
    if text == "A": return "1"
    if text == "B": return "2"
    if text == "C": return "3"

    # Embedded digit  "SHIFT-1", "SHIFT 2", "SHIFT1"
    for ch in text:
        if ch in ("1", "2", "3"):
            return ch

    return ""


# =====================================================
# GET ALL EXCEL FILES
# =====================================================

def get_excel_files():
    excel_files = []
    for root, dirs, files in os.walk(DATABASE_ROOT):
        for file in files:
            if file.startswith("~$"):
                continue
            if file.lower().endswith((".xlsx", ".xlsm")):
                excel_files.append(os.path.join(root, file))
    return excel_files


# =====================================================
# PARSE A RAW DATE CELL  (col 10 in DB)
# Returns a datetime or None.
# =====================================================

def parse_date_cell(raw):
    """
    The DATE column (index 10) in the DB can hold:
      • a datetime object    (openpyxl native)
      • a date object
      • a string in various formats
    Returns datetime or None.
    """
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw
    # openpyxl may return a date (not datetime) object
    if hasattr(raw, "year") and hasattr(raw, "month") and hasattr(raw, "day"):
        try:
            return datetime(raw.year, raw.month, raw.day)
        except Exception:
            pass
    raw_str = str(raw).strip()
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%d.%m.%Y",
        "%d.%m.%y",
    ):
        try:
            return datetime.strptime(raw_str, fmt)
        except Exception:
            pass
    return None


# =====================================================
# MAIN DASHBOARD DATA
#
# DB column layout (0-based):
#   0  Year
#   1  Month (e.g. "Jun")
#   2  Product
#   3  Length
#   4  Product Type
#   5  Link (file path)
#   6  Sheet Name
#   7  MID
#   8  ORDER NO
#   9  Team Leader
#   10 DATE  ← raw date cell from sheet
#   11 SHIFT (int 1/2/3 or raw value)
#   12 CLOCK ID
#   13 Product Name
#   14 Line Name   ← ws K2
#   15 Total Produced
#   16 OK
#   17 NOK
# =====================================================

def get_dashboard_data(
    selected_product="ALL",
    selected_date="",
    selected_shift="ALL"
):
    product_counts = {
        "OPTITAP": 0,
        "BTCOF":   0,
        "PRODIGY": 0,
        "JUMPER":  0,
        "1x4":     0,
        "1x8":     0
    }

    summary = {}
    excel_files = get_excel_files()

    # ------------------------------------
    # PRODUCT FILTER
    # ------------------------------------
    selected_products = []
    if selected_product and selected_product != "ALL":
        selected_products = [p.strip() for p in selected_product.split(",")]

    # ------------------------------------
    # DATE FILTER
    # FIX: parse the selected date once; then match against
    #      col 10 (DATE) using parse_date_cell().
    #      Previously the code also tried to match Year (col 0)
    #      and Month abbr (col 1) separately — that is kept as a
    #      fast pre-check, but the definitive test is col 10.
    # ------------------------------------
    selected_year  = None
    selected_month = None
    selected_day   = None

    if selected_date:
        try:
            dt = datetime.strptime(selected_date, "%Y-%m-%d")
            selected_year  = dt.year
            selected_month = dt.strftime("%b").upper()   # e.g. "JUN"
            selected_day   = dt.day
        except Exception:
            pass

    # ------------------------------------
    # SHIFT FILTER
    # "ALL" passes everything; otherwise normalise to "1"/"2"/"3"
    # ------------------------------------
    filter_shift = "ALL"
    if selected_shift and selected_shift != "ALL":
        filter_shift = normalize_shift(selected_shift)

    # ------------------------------------
    # LOOP DB FILES
    # ------------------------------------
    for file_path in excel_files:
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    year         = row[0]                  # col A – Year  (int)
                    month        = safe_str(row[1]).upper() # col B – Month abbr
                    product      = safe_str(row[2])         # col C – Product
                    length       = safe_str(row[3])         # col D – Length
                    mid          = safe_str(row[7])         # col H – MID
                    order_no     = safe_str(row[8])         # col I – Order No
                    raw_date_cell = row[10]                 # col K – DATE
                    shift        = normalize_shift(row[11]) # col L – SHIFT
                    line         = safe_str(row[14])        # col O – Line Name

                    # FIX: use col 15 (Total Produced) not col 15 ambiguously
                    produced_raw = row[15]
                    produced     = int(produced_raw) if produced_raw is not None else 0

                except Exception:
                    continue

                # Skip completely empty rows
                if produced == 0 and not product:
                    continue

                # ------------------------------------
                # DATE FILTER
                # FIX: Use a two-step check:
                #   1. Quick reject on year+month (cheap).
                #   2. Definitive day-level check against col 10.
                #      If col 10 is missing/unparseable fall back
                #      to year+month match only (don't silently
                #      pass everything through as before).
                # ------------------------------------
                if selected_date:
                    # Quick year+month pre-check
                    if year != selected_year:
                        continue
                    if month != selected_month:
                        continue

                    # Day-level check via the actual DATE cell
                    parsed_date = parse_date_cell(raw_date_cell)
                    if parsed_date is not None:
                        if parsed_date.day != selected_day:
                            continue
                    # If date cell is unparseable, year+month match is
                    # accepted as best-effort (don't silently include all)

                # ------------------------------------
                # PRODUCT FILTER
                # ------------------------------------
                if selected_products:
                    if product not in selected_products:
                        continue

                # ------------------------------------
                # SHIFT FILTER
                # ------------------------------------
                if filter_shift != "ALL":
                    if shift != filter_shift:
                        continue

                # ------------------------------------
                # PRODUCT COUNTS
                # ------------------------------------
                if product in product_counts:
                    product_counts[product] += produced

                # ------------------------------------
                # TABLE SUMMARY
                # Only show rows with a valid MID format
                # ------------------------------------
                display_mid = mid if is_valid_mid(mid) else ""
                key = (display_mid, order_no, product, length, line)

                if key not in summary:
                    summary[key] = {
                        "mid":     display_mid,
                        "order":   order_no,
                        "product": product,
                        "length":  length,
                        "line":    line,
                        "shift_a": 0,
                        "shift_b": 0,
                        "shift_c": 0,
                        "total":   0
                    }

                if shift == "1":
                    summary[key]["shift_a"] += produced
                elif shift == "2":
                    summary[key]["shift_b"] += produced
                elif shift == "3":
                    summary[key]["shift_c"] += produced

                summary[key]["total"] += produced

        except Exception as e:
            print(f"Workbook Error: {file_path}")
            print(e)

    # ------------------------------------
    # CONVERT TO LIST
    # ------------------------------------
    rows = list(summary.values())
    rows.sort(key=lambda x: (str(x["line"]), str(x["mid"])))

    total = sum(product_counts.values())

    return {
        "OPTITAP": product_counts["OPTITAP"],
        "BTCOF":   product_counts["BTCOF"],
        "PRODIGY": product_counts["PRODIGY"],
        "JUMPER":  product_counts["JUMPER"],
        "1x4":     product_counts["1x4"],
        "1x8":     product_counts["1x8"],
        "total":   total,
        "rows":    rows
    }


# =====================================================
# TEST
# =====================================================

if __name__ == "__main__":
    data = get_dashboard_data(
        selected_product="ALL",
        selected_date="",
        selected_shift="ALL"
    )

    print("\nPRODUCT COUNTS")
    print("-" * 50)
    print("OPTITAP :", data["OPTITAP"])
    print("BTCOF   :", data["BTCOF"])
    print("PRODIGY :", data["PRODIGY"])
    print("JUMPER  :", data["JUMPER"])
    print("1x4     :", data["1x4"])
    print("1x8     :", data["1x8"])
    print("TOTAL   :", data["total"])

    print("\nMID SUMMARY")
    print("-" * 100)
    for row in data["rows"]:
        print(
            row["mid"], row["order"], row["product"],
            row["length"], row["line"],
            row["shift_a"], row["shift_b"], row["shift_c"], row["total"]
        )