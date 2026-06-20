import os
from datetime import datetime
from flask import Flask, render_template, jsonify, request

from processor import get_dashboard_data
from consolidator import process_files, get_available_dates
from watcher import start_watcher, watcher_status

app = Flask(__name__)

start_watcher()

# ── Excel layout constants (Efficiency sheet, no hourly rows) ─────────────────
EFF_LINES           = ["Line1", "Line2", "Line3", "Line4"]
EFF_SHIFT_START_COL = {"A": 2, "B": 7, "C": 12}

ROW_PRODUCT = 3
ROW_TARGET  = 4
ROW_ACTUAL  = 5
ROW_EFF     = 6   # formula =Actual/Target


def _excel_path():
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), "Data_sheet.xlsx")


def _load_targets():
    from openpyxl import load_workbook
    defaults = {
        "OPTITAP": 350, "PRODIGY": 350,
        "BTCOF": 350, "BT COF": 350, "BT-COF": 350,
        "JUMPER": 175,
        "1X4": 350, "1X8": 350,
    }
    try:
        wb = load_workbook(_excel_path(), data_only=True)
        ws = wb["Targets"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                key = str(row[0]).strip().upper()
                val = int(row[1])
                defaults[key] = val
                defaults[key.replace(" ", "").replace("-", "")] = val
        wb.close()
    except Exception:
        pass
    return defaults


def _push_to_excel(records):
    """
    Write Product, Actual to the Efficiency sheet.
    records: [{line, shift, product, actual}]
    Efficiency % is an Excel formula — recalcs automatically.
    """
    from openpyxl import load_workbook
    wb = load_workbook(_excel_path())
    ws = wb["Efficiency"]

    # First write targets for each shift/line combo (from Targets sheet)
    targets = _load_targets()

    written = 0
    for rec in records:
        line    = rec.get("line", "")
        shift   = rec.get("shift", "").upper()
        product = rec.get("product", "")
        actual  = rec.get("actual", 0)

        if line not in EFF_LINES or shift not in EFF_SHIFT_START_COL:
            continue

        col = EFF_SHIFT_START_COL[shift] + EFF_LINES.index(line)
        lookup = product.upper().replace(" ", "").replace("-", "")
        target_val = targets.get(lookup, targets.get(product.upper(), 350))

        ws.cell(row=ROW_PRODUCT, column=col, value=product)
        ws.cell(row=ROW_TARGET,  column=col, value=target_val)
        ws.cell(row=ROW_ACTUAL,  column=col, value=int(actual))
        written += 1

    wb.save(_excel_path())
    return written


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route("/")
def home():
    return render_template("dashboard.html")


@app.route("/dashboard-data")
def dashboard_data():
    product = request.args.get("product", "ALL")
    date    = request.args.get("date",    "")
    shift   = request.args.get("shift",   "ALL")
    try:
        data = get_dashboard_data(product, date, shift)
        return jsonify(data)
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/watcher-status")
def get_watcher_status():
    return jsonify(watcher_status)


@app.route("/get-available-dates")
def available_dates():
    try:
        dates = get_available_dates()
        return jsonify({"success": True, "dates": dates})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/refresh-database")
def refresh_database():
    dates_param  = request.args.get("dates", "")
    filter_dates = None
    if dates_param.strip():
        filter_dates = [d.strip() for d in dates_param.split(",") if d.strip()]
    try:
        result = process_files(filter_dates=filter_dates)
        return jsonify({
            "success": True,
            "message": f"Refresh Completed\nFiles Scanned : {result['total_files']}",
            "result":  result
        })
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/health")
def health():
    return jsonify({"status": "Running"})


# ── Efficiency Data ────────────────────────────────────────────────────────────
# Returns line-wise actual vs target per shift. No hourly breakdown.
# Auto-pushes to Excel after loading.
@app.route("/efficiency-data")
def efficiency_data():
    try:
        from processor import normalize_shift, get_excel_files, safe_str, parse_date_cell
        from openpyxl import load_workbook

        subproduct = request.args.get("subproduct", "ALL")
        shift_arg  = request.args.get("shift",      "ALL")
        date_arg   = request.args.get("date",        "")

        active_shifts = (
            ["A", "B", "C"] if shift_arg == "ALL"
            else ([shift_arg] if shift_arg in ("A", "B", "C") else ["A", "B", "C"])
        )

        # Parse date filter
        sel_year = sel_month = sel_day = None
        if date_arg:
            try:
                dt        = datetime.strptime(date_arg, "%Y-%m-%d")
                sel_year  = dt.year
                sel_month = dt.strftime("%b").upper()
                sel_day   = dt.day
            except Exception:
                pass

        # Aggregate totals per line + shift + product
        line_totals = {}   # {line: {shift_letter: {product: set(), actual: int}}}

        for file_path in get_excel_files():
            try:
                wb2 = load_workbook(file_path, data_only=True)
                ws  = wb2.active
                for row in ws.iter_rows(min_row=2, values_only=True):
                    try:
                        year         = row[0]
                        month        = safe_str(row[1]).upper()
                        product      = safe_str(row[2])
                        raw_date     = row[10]
                        shift_num    = normalize_shift(row[11])
                        line         = safe_str(row[14])
                        produced_raw = row[15]
                        produced     = int(produced_raw) if produced_raw is not None else 0

                        if not line or produced == 0:
                            continue
                        # Normalize: "3" -> "Line3", "line3" -> "Line3", "Line3" stays
                        if line.isdigit():
                            line = f"Line{line}"
                        elif line.lower().startswith("line"):
                            line = "Line" + line[4:].strip()
                        else:
                            continue  # skip unrecognised line names

                        if date_arg:
                            if year != sel_year or month != sel_month:
                                continue
                            parsed = parse_date_cell(raw_date)
                            if parsed is not None and parsed.day != sel_day:
                                continue

                        if subproduct != "ALL" and product != subproduct:
                            continue

                        shift_letter = {"1": "A", "2": "B", "3": "C"}.get(str(shift_num), "")
                        if not shift_letter:
                            continue
                        if shift_arg != "ALL" and shift_letter != shift_arg:
                            continue

                        if line not in line_totals:
                            line_totals[line] = {}
                        if shift_letter not in line_totals[line]:
                            line_totals[line][shift_letter] = {"actual": 0, "product": ""}
                        line_totals[line][shift_letter]["actual"] += produced
                        if product:
                            line_totals[line][shift_letter]["product"] = product

                    except Exception:
                        continue
                wb2.close()
            except Exception:
                continue

        # Load targets
        targets    = _load_targets()
        lookup_key = subproduct.upper().replace(" ", "").replace("-", "")
        target_val = (
            targets.get(lookup_key, targets.get(subproduct.upper(), 350))
            if subproduct != "ALL" else 350
        )

        # Build result
        result = {}
        for line, shifts in line_totals.items():
            result[line] = {}
            for sh, d in shifts.items():
                actual = d["actual"]
                result[line][sh] = {
                    "actual":         actual,
                    "target":         target_val,
                    "efficiency_pct": round((actual / target_val) * 100, 1) if target_val else 0,
                    "product":        d["product"],
                }

        # Auto-push to Excel
        excel_written = 0
        push_error    = None
        try:
            records = []
            for line, shifts in result.items():
                for sh, d in shifts.items():
                    records.append({
                        "line":    line,
                        "shift":   sh,
                        "product": d["product"],
                        "actual":  d["actual"],
                    })
            excel_written = _push_to_excel(records)
        except Exception as e:
            push_error = str(e)

        return jsonify({
            "success":       True,
            "data":          result,
            "target":        target_val,
            "active_shifts": active_shifts,
            "lines":         sorted(result.keys()),
            "excel_written": excel_written,
            "excel_error":   push_error,
            "updated_at":    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        })

    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/health")
def health_check():
    return jsonify({"status": "Running"})


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=1111, debug=True, use_reloader=False)