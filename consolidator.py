import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

# =====================================================
# CONFIGURATION
# =====================================================

SOURCE_FOLDER = r"\\atgv1pfss03\Groups\Enterprise Products\BBN\HFF EOL\2026\New 2026\New 2026"
OUTPUT_FOLDER = r"C:\Users\IR1035\Desktop\Database"

PRODUCT_GROUPS = {
    "OPTITAP": "HFF",
    "PRODIGY": "HFF",
    "JUMPER":  "HFF",
    "BTCOF":   "HFF",
    "1x4":     "FST",
    "1x8":     "FST"
}

PRODUCT_DISPLAY_NAMES = {
    "OPTITAP": "OPTITAP",
    "PRODIGY": "PRODIGY",
    "JUMPER":  "JUMPER 1M",
    "BTCOF":   "BT-COF-256",
    "1x4":     "1x4",
    "1x8":     "1x8"
}


# =====================================================
# FILE DATE HELPERS
# =====================================================

def get_file_modified_date(file_path):
    """Returns date of last modification — used for filtering."""
    try:
        timestamp = os.path.getmtime(file_path)
        return datetime.fromtimestamp(timestamp).date()
    except Exception:
        return None


def normalize_path(path):
    """
    Normalise a file path so comparisons work regardless of
    drive-letter case, forward/back slash, or trailing separator.
    """
    return os.path.normcase(os.path.normpath(str(path)))


# =====================================================
# SCAN SOURCE FOLDER — return all unique modified dates
# =====================================================

def get_available_dates():
    """Returns sorted list of unique file-modified dates."""
    dates = set()
    for root, dirs, files in os.walk(SOURCE_FOLDER):
        for file in files:
            if file.startswith("~$"):
                continue
            if not file.lower().endswith((".xlsx", ".xlsm")):
                continue
            file_path = os.path.join(root, file)
            modified_date = get_file_modified_date(file_path)
            if modified_date:
                dates.add(str(modified_date))
    return sorted(dates)


# =====================================================
# DATE PARSER  (for E2 cell inside sheets)
# =====================================================

def get_sheet_date(value):
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        value = value.strip()
        formats = [
            "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y",
            "%m/%d/%Y", "%d.%m.%Y", "%d.%m.%y",
        ]
        for fmt in formats:
            try:
                return datetime.strptime(value, fmt)
            except Exception:
                pass
    return None


# =====================================================
# SHIFT NORMALISER
# FIX: always return a plain int (1, 2, or 3) or None.
#      processor.py normalises to strings on its own side.
# =====================================================

def normalize_shift(raw):
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        v = int(raw)
        return v if v in (1, 2, 3) else None
    text = str(raw).strip().upper()
    if text in ("1", "2", "3"):
        return int(text)
    # Ordinal: "1ST", "2ND", "3RD"
    if text.startswith("1"): return 1
    if text.startswith("2"): return 2
    if text.startswith("3"): return 3
    # Named shifts
    if text == "A": return 1
    if text == "B": return 2
    if text == "C": return 3
    # Embedded digit: "SHIFT-1", "SHIFT 2", "SHIFT1"
    for ch in text:
        if ch in ("1", "2", "3"):
            return int(ch)
    return None


# =====================================================
# PRODUCT DETECTION
# =====================================================

def detect_product(ws):
    cells_to_check = [
        "A1", "B1", "C1", "D1", "E1",
        "G2", "H2", "I2",
        "H3", "I3"
    ]
    for cell in cells_to_check:
        try:
            value = ws[cell].value
            if not value:
                continue
            text = str(value).upper()
            if "OPTITAP"  in text: return "OPTITAP"
            if "PRODIGY"  in text: return "PRODIGY"
            if "JUMPER"   in text: return "JUMPER"
            if "BT-COF"   in text: return "BTCOF"
            if "COF"      in text: return "BTCOF"
            if "1X8"      in text: return "1x8"
            if "1X4"      in text: return "1x4"
            if "SPLITTER" in text: return "1x4"
        except Exception:
            pass
    return None


# =====================================================
# OK / NOK COUNTING
#
# FIX — previous version counted ANY non-empty cell in
# rows 16-27 as a result, even numeric measurements,
# which inflated the produced count and mislabelled
# units as NOK.
#
# Rule now:
#   • A column is only counted if row 8 contains a
#     non-empty serial number (unchanged).
#   • Within rows 16-27, ONLY look at cells that
#     contain a recognised pass/fail keyword.
#     Numeric values and measurement strings are ignored.
#   • If a column has at least one FAIL/NOK cell it is NOK.
#   • If all recognised verdict cells say PASS/OK it is OK.
#   • If NO verdict cell is found, the unit is considered
#     PRODUCED but verdict is treated as NOK (incomplete).
# =====================================================

def calculate_ok_nok(ws):
    SERIAL_ROW = 8
    START_ROW  = 16
    END_ROW    = 27
    START_COL  = 7   # Column G

    PASS_KEYWORDS = {"PASS", "OK"}
    FAIL_KEYWORDS = {"FAIL", "NOK", "NOT OK", "NG"}

    produced = 0
    ok       = 0
    nok      = 0

    for col in range(START_COL, ws.max_column + 1):

        serial_no = ws.cell(row=SERIAL_ROW, column=col).value
        if serial_no is None or str(serial_no).strip() == "":
            continue   # no serial → not a test column

        produced += 1

        has_fail  = False
        has_pass  = False

        for row in range(START_ROW, END_ROW + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value is None:
                continue

            # Only evaluate string cells; skip numbers (measurements)
            if not isinstance(cell_value, str):
                continue

            text = cell_value.strip().upper()
            if text == "":
                continue

            if text in PASS_KEYWORDS:
                has_pass = True
            elif text in FAIL_KEYWORDS:
                has_fail = True
            # Any other string (unexpected label) counts as fail
            else:
                has_fail = True

        if has_pass and not has_fail:
            ok += 1
        else:
            # Covers: explicit fail, mixed, or no verdict found
            nok += 1

    return produced, ok, nok


# =====================================================
# EXTRACT DATA FROM ONE WORKBOOK
# =====================================================

def extract_workbook(file_path):
    results = []
    try:
        wb = load_workbook(file_path, data_only=True)

        for sheet_name in wb.sheetnames:
            if "REV" in sheet_name.upper():
                continue

            ws = wb[sheet_name]

            raw_date = ws["E2"].value
            if not raw_date:
                continue

            file_date = get_sheet_date(raw_date)
            if not file_date:
                print(f"  [WARN] Could not parse date '{raw_date}' in '{sheet_name}' — skipping.")
                continue

            product = detect_product(ws)
            if not product:
                continue

            mid_value = ws["B2"].value
            if not mid_value:
                mid_value = os.path.splitext(os.path.basename(file_path))[0]

            length = sheet_name.strip()

            produced, ok, nok = calculate_ok_nok(ws)
            print(f"  {sheet_name} → Produced={produced}, OK={ok}, NOK={nok}")

            if produced == 0:
                continue

            raw_shift    = ws["E3"].value
            norm_shift   = normalize_shift(raw_shift)
            # Store as int (1/2/3) when recognised, raw value otherwise
            stored_shift = norm_shift if norm_shift is not None else raw_shift

            # FIX: col indices match the DB header order exactly:
            #   0=Year  1=Month  2=Product  3=Length  4=ProductType
            #   5=Link  6=SheetName  7=MID  8=OrderNo  9=TeamLeader
            #   10=DATE  11=SHIFT  12=ClockID  13=ProductName  14=LineName
            #   15=TotalProduced  16=OK  17=NOK
            row_data = [
                file_date.year,               # 0  Year
                file_date.strftime("%b"),      # 1  Month
                product,                       # 2  Product
                length,                        # 3  Length (sheet tab name)
                PRODUCT_GROUPS.get(product, "UNKNOWN"),  # 4  Product Type
                normalize_path(file_path),     # 5  Link  ← normalised path
                sheet_name,                    # 6  Sheet Name
                mid_value,                     # 7  MID       (ws B2)
                ws["B3"].value,                # 8  Order No  (ws B3)
                ws["B4"].value,                # 9  Team Leader (ws B4)
                file_date,                     # 10 DATE
                stored_shift,                  # 11 SHIFT
                ws["E4"].value,                # 12 Clock ID  (ws E4)
                ws["G2"].value,                # 13 Product Name (ws G2)
                ws["K2"].value,                # 14 Line Name  (ws K2)
                produced,                      # 15 Total Produced
                ok,                            # 16 OK
                nok,                           # 17 NOK
            ]
            results.append(row_data)

    except Exception as e:
        print(f"\nERROR: {file_path}")
        print(e)

    return results


# =====================================================
# PROCESS SPECIFIC FILES  (used by watcher)
#
# FIX: path comparison now uses normalize_path() so UNC
#      paths with mixed separators still deduplicate
#      correctly and old rows are never kept alongside
#      freshly written ones.
# =====================================================

def process_specific_files(file_paths):
    """
    Called by watcher with a list of changed file paths.
    Re-extracts each file and REPLACES its rows in the database.
    """
    headers = [
        "Year", "Month", "Product", "Length", "Product Type",
        "Link", "Sheet Name",
        "MID", "ORDER NO", "Team Leader",
        "DATE", "SHIFT", "CLOCK ID",
        "Product Name", "Line Name",
        "Total Produced", "OK", "NOK"
    ]

    # Normalise paths so comparisons are case/separator agnostic
    updating_norm = {normalize_path(p) for p in file_paths}

    # Extract fresh data from each changed file
    new_rows_by_key = {}   # (year, month, product) → [rows]
    for file_path in file_paths:
        print(f"  [WATCHER] Re-extracting: {os.path.basename(file_path)}")
        rows = extract_workbook(file_path)
        for row in rows:
            key = (row[0], row[1], row[2])   # year, month, product
            new_rows_by_key.setdefault(key, []).append(row)

    if not new_rows_by_key:
        print("  [WATCHER] No data extracted from changed files.")
        return 0

    files_updated = 0

    for (year, month, product), new_rows in new_rows_by_key.items():
        product_type = PRODUCT_GROUPS.get(product, "UNKNOWN")
        display_name = PRODUCT_DISPLAY_NAMES.get(product, product)

        folder_path = os.path.join(OUTPUT_FOLDER, str(year), month, product_type)
        os.makedirs(folder_path, exist_ok=True)
        output_file = os.path.join(folder_path, f"{display_name}.xlsx")

        # Load existing DB, drop rows that came from the changed files
        kept_rows = []
        if os.path.exists(output_file):
            try:
                wb_existing = load_workbook(output_file, data_only=True)
                ws_existing = wb_existing.active
                for existing_row in ws_existing.iter_rows(min_row=2, values_only=True):
                    existing_row = list(existing_row)
                    # col 5 = Link (normalised path stored by extract_workbook)
                    row_norm = normalize_path(existing_row[5]) if existing_row[5] else ""
                    if row_norm not in updating_norm:
                        kept_rows.append(existing_row)
                wb_existing.close()
            except Exception as e:
                print(f"  [WATCHER] Could not read existing DB file: {e}")
                kept_rows = []

        all_rows = kept_rows + new_rows

        wb_out       = Workbook()
        ws_out       = wb_out.active
        ws_out.title = display_name
        ws_out.append(headers)
        for row in all_rows:
            ws_out.append(row)

        wb_out.save(output_file)
        files_updated += 1
        print(f"  [WATCHER] DB updated: {output_file}  ({len(new_rows)} fresh rows, {len(kept_rows)} kept)")

    return files_updated


# =====================================================
# MAIN PROCESS  (manual refresh from dashboard)
# =====================================================

def process_files(filter_dates=None):
    """
    Manual full or date-filtered consolidation.
    filter_dates: list of 'YYYY-MM-DD' strings (filters by mtime).
                  None = full refresh.
    """
    date_filter_set = None
    if filter_dates:
        date_filter_set = set()
        for d in filter_dates:
            try:
                date_filter_set.add(datetime.strptime(d, "%Y-%m-%d").date())
            except Exception:
                pass

    headers = [
        "Year", "Month", "Product", "Length", "Product Type",
        "Link", "Sheet Name",
        "MID", "ORDER NO", "Team Leader",
        "DATE", "SHIFT", "CLOCK ID",
        "Product Name", "Line Name",
        "Total Produced", "OK", "NOK"
    ]

    grouped_data  = {}
    total_files   = 0
    skipped_files = 0

    print("=" * 80)
    if date_filter_set:
        print(f"CONSOLIDATION  (dates: {sorted(str(d) for d in date_filter_set)})")
    else:
        print("CONSOLIDATION  (full refresh)")
    print("=" * 80)

    for root, dirs, files in os.walk(SOURCE_FOLDER):
        for file in files:
            if file.startswith("~$"):
                continue
            if not file.lower().endswith((".xlsx", ".xlsm")):
                continue

            file_path     = os.path.join(root, file)
            modified_date = get_file_modified_date(file_path)

            if date_filter_set:
                if modified_date not in date_filter_set:
                    skipped_files += 1
                    print(f"  SKIP  [{modified_date}]  {file}")
                    continue

            total_files += 1
            print(f"  READ  [{modified_date}]  {file}")
            rows = extract_workbook(file_path)

            for row in rows:
                key = (row[0], row[1], row[2])
                grouped_data.setdefault(key, []).append(row)

    print(f"\nFiles read   : {total_files}")
    print(f"Files skipped: {skipped_files}")
    print("\nSaving to Database...\n")

    files_created = 0

    for (year, month, product), rows in grouped_data.items():
        product_type = PRODUCT_GROUPS.get(product, "UNKNOWN")
        display_name = PRODUCT_DISPLAY_NAMES.get(product, product)

        folder_path = os.path.join(OUTPUT_FOLDER, str(year), month, product_type)
        os.makedirs(folder_path, exist_ok=True)
        output_file = os.path.join(folder_path, f"{display_name}.xlsx")

        # For date-filtered refresh: merge with other-date rows
        if os.path.exists(output_file) and date_filter_set:
            try:
                wb_existing   = load_workbook(output_file, data_only=True)
                ws_existing   = wb_existing.active
                existing_rows = [list(r) for r in ws_existing.iter_rows(min_row=2, values_only=True)]
                wb_existing.close()
            except Exception:
                existing_rows = []

            # FIX: deduplicate by normalised path so mixed-separator
            #      paths from UNC shares don't survive as duplicates
            updating_norm = {normalize_path(r[5]) for r in rows if r[5]}
            kept_rows = [
                r for r in existing_rows
                if normalize_path(r[5]) not in updating_norm
            ]
            all_rows = kept_rows + rows
        else:
            all_rows = rows

        wb_out       = Workbook()
        ws_out       = wb_out.active
        ws_out.title = display_name
        ws_out.append(headers)
        for row in all_rows:
            ws_out.append(row)

        wb_out.save(output_file)
        files_created += 1
        print(f"  Saved: {output_file}  ({len(all_rows)} rows)")

    print("\n" + "=" * 80)
    print("COMPLETED")
    print("=" * 80)

    return {
        "total_files":   total_files,
        "skipped_files": skipped_files,
        "files_created": files_created
    }


if __name__ == "__main__":
    process_files()